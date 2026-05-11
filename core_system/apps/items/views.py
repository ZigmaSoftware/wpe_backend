from __future__ import annotations

import re
from typing import Any

from django.db.models import ProtectedError
from openpyxl import load_workbook
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated

from common.drf import QueryParamFilterMixin, StandardResultsSetPagination, success_response

from .models import Item
from .serializers import ItemSerializer


HEADER_ALIASES = {
    "product_type": "product_type",
    "product type": "product_type",
    "item_type": "product_type",
    "item type": "product_type",
    "category": "category",
    "group": "group",
    "sub_group": "sub_group",
    "subgroup": "sub_group",
    "sub group": "sub_group",
    "item_name": "item_name",
    "item name": "item_name",
    "external_item_id": "external_item_id",
    "external item id": "external_item_id",
    "item_code": "item_code",
    "item code": "item_code",
    "hsn_code": "hsn_code",
    "hsn code": "hsn_code",
    "hsn": "hsn_code",
    "unit": "unit",
    "product_details": "product_details",
    "product details": "product_details",
    "description": "description",
    "min_max_status": "min_max_status",
    "min max status": "min_max_status",
    "status": "status",
}
REQUIRED_FIELDS = ("category", "group", "sub_group", "item_name", "unit")
BOOLEAN_FIELDS = {"min_max_status", "status"}
TEXT_DEFAULTS = {
    "product_type": Item.PRODUCT_TYPE_GENERAL,
    "hsn_code": None,
    "product_details": None,
    "description": None,
    "external_item_id": None,
}
BOOLEAN_DEFAULTS = {
    "min_max_status": False,
    "status": True,
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def is_blank_row(row: tuple[Any, ...]) -> bool:
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str) and not cell.strip():
            continue
        return False
    return True


def stringify_cell(value: Any, *, allow_blank: bool = False) -> str | None:
    if value is None:
        return None if allow_blank else ""

    if isinstance(value, bool):
        text = "true" if value else "false"
    elif isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        text = str(int(value)) if value.is_integer() else str(value)
    else:
        text = str(value).strip()

    if not text:
        return None if allow_blank else ""
    return text


def parse_boolean(value: Any, *, default: bool) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(int(value))

    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "on", "active", "enabled"}:
        return True
    if normalized in {"0", "false", "no", "n", "off", "inactive", "disabled"}:
        return False
    raise ValueError(f"Cannot interpret boolean value: {value}")


def format_serializer_errors(errors: dict[str, Any]) -> str:
    parts: list[str] = []
    for field, messages in errors.items():
        if isinstance(messages, (list, tuple)):
            message_text = ", ".join(str(message) for message in messages)
        else:
            message_text = str(messages)
        parts.append(f"{field}: {message_text}")
    return "; ".join(parts) if parts else "Invalid item row"


def resolve_existing_item(payload: dict[str, Any]) -> Item | None:
    external_item_id = payload.get("external_item_id")
    if external_item_id:
        existing_item = Item.objects.filter(external_item_id__iexact=external_item_id).first()
        if existing_item is not None:
            return existing_item

    return Item.objects.filter(
        item_name=payload["item_name"],
        category=payload["category"],
        group=payload["group"],
        sub_group=payload["sub_group"],
        unit=payload["unit"],
    ).first()


class ItemViewSet(QueryParamFilterMixin, viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, FormParser, MultiPartParser]
    queryset = Item.objects.all().order_by("-id")
    serializer_class = ItemSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = (
        "item_code",
        "item_name",
        "external_item_id",
        "category",
        "group",
        "sub_group",
        "product_type",
        "unit",
    )
    ordering_fields = ("created_at", "updated_at", "item_name", "item_code", "id")
    filterset_map = {
        "product_type": "product_type",
        "category": "category",
        "group": "group",
        "sub_group": "sub_group",
        "status": "status",
    }

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return success_response(
                message="Items fetched successfully.",
                data=self.paginator.get_paginated_data(serializer.data),
            )

        serializer = self.get_serializer(queryset, many=True)
        return success_response(
            message="Items fetched successfully.",
            data={"count": len(serializer.data), "results": serializer.data},
        )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return success_response(
            message="Item created successfully.",
            data=self.get_serializer(serializer.instance).data,
            status_code=status.HTTP_201_CREATED,
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return success_response(
            message="Item updated successfully.",
            data=self.get_serializer(serializer.instance).data,
        )

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        return success_response(
            message="Item fetched successfully.",
            data=self.get_serializer(instance).data,
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
        except ProtectedError:
            return success_response(
                message="Item cannot be deleted because dependent records exist.",
                data={},
                status_code=status.HTTP_400_BAD_REQUEST,
                success=False,
            )
        return success_response(message="Item deleted successfully.")

    @action(detail=False, methods=["post"], url_path="import")
    def import_excel(self, request):
        uploaded_file = request.FILES.get("file")
        if uploaded_file is None:
            return success_response(
                message="Please upload an Excel file using the 'file' field.",
                data={},
                status_code=status.HTTP_400_BAD_REQUEST,
                success=False,
            )

        if not uploaded_file.name.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            return success_response(
                message="Only .xlsx Excel files are supported.",
                data={},
                status_code=status.HTTP_400_BAD_REQUEST,
                success=False,
            )

        try:
            workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
        except Exception:
            return success_response(
                message="The uploaded file is not a valid Excel workbook.",
                data={},
                status_code=status.HTTP_400_BAD_REQUEST,
                success=False,
            )

        try:
            sheet = workbook.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return success_response(
                    message="The workbook does not contain a header row.",
                    data={},
                    status_code=status.HTTP_400_BAD_REQUEST,
                    success=False,
                )

            column_map: dict[int, str] = {}
            for index, header in enumerate(header_row):
                field_name = HEADER_ALIASES.get(normalize_header(header))
                if field_name and field_name not in column_map.values():
                    column_map[index] = field_name

            missing_required_fields = [field for field in REQUIRED_FIELDS if field not in column_map.values()]
            if missing_required_fields:
                return success_response(
                    message="The workbook is missing required columns: " + ", ".join(missing_required_fields),
                    data={},
                    status_code=status.HTTP_400_BAD_REQUEST,
                    success=False,
                )

            created_count = 0
            updated_count = 0
            failed_rows: list[dict[str, Any]] = []
            processed_count = 0

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if is_blank_row(row):
                    continue

                processed_count += 1
                payload: dict[str, Any] = {**TEXT_DEFAULTS, **BOOLEAN_DEFAULTS}

                try:
                    for column_index, field_name in column_map.items():
                        cell_value = row[column_index] if column_index < len(row) else None
                        if field_name in BOOLEAN_FIELDS:
                            payload[field_name] = parse_boolean(cell_value, default=BOOLEAN_DEFAULTS[field_name])
                        elif field_name in TEXT_DEFAULTS:
                            payload[field_name] = stringify_cell(cell_value, allow_blank=True)
                        else:
                            payload[field_name] = stringify_cell(cell_value)
                except ValueError as exc:
                    failed_rows.append({"row": row_number, "message": str(exc)})
                    continue

                existing_item = resolve_existing_item(payload)
                serializer = ItemSerializer(instance=existing_item, data=payload, partial=existing_item is not None)
                if not serializer.is_valid():
                    failed_rows.append(
                        {
                            "row": row_number,
                            "message": format_serializer_errors(serializer.errors),
                            "details": serializer.errors,
                        }
                    )
                    continue

                serializer.save()
                if existing_item is None:
                    created_count += 1
                else:
                    updated_count += 1

            failed_count = len(failed_rows)
            response_body = {
                "created_count": created_count,
                "updated_count": updated_count,
                "failed_count": failed_count,
                "processed_count": processed_count,
                "errors": failed_rows,
            }

            if created_count == 0 and updated_count == 0:
                return success_response(
                    message=(
                        f"All rows failed to import. First error: row {failed_rows[0]['row']}: "
                        f"{failed_rows[0]['message']}"
                        if failed_rows
                        else "The workbook does not contain any item rows."
                    ),
                    data=response_body,
                    status_code=status.HTTP_400_BAD_REQUEST,
                    success=False,
                )

            status_code = status.HTTP_207_MULTI_STATUS if failed_count > 0 else status.HTTP_201_CREATED
            return success_response(
                message="Item import processed successfully.",
                data=response_body,
                status_code=status_code,
            )
        finally:
            workbook.close()
