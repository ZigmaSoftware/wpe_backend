from __future__ import annotations

import re
from copy import deepcopy
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any
from urllib.parse import urlencode

from django.db import IntegrityError, OperationalError, ProgrammingError, transaction
from django.db.models import Q
from django.forms.models import model_to_dict
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from rest_framework import status, viewsets
from rest_framework.exceptions import ParseError, UnsupportedMediaType, ValidationError as DRFValidationError
from rest_framework.permissions import AllowAny
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.store.services import add_stock_from_grn
from .models import GRN, GRNAuditLog, QCR
from .serializers import GRNAuditLogSerializer, GRNReadSerializer, GRNSerializer, QCRSerializer


HEADER_ALIASES = {
    "po_no": "po_no",
    "po number": "po_no",
    "po no": "po_no",
    "po": "po_no",
    "po_date": "po_date",
    "po date": "po_date",
    "grn_no": "grn_no",
    "grn number": "grn_no",
    "grn no": "grn_no",
    "grn": "grn_no",
    "grn_date": "grn_date",
    "grn date": "grn_date",
    "supplier_invoice_no": "supplier_invoice_no",
    "supplier invoice no": "supplier_invoice_no",
    "supplier invoice number": "supplier_invoice_no",
    "supplier_invoice_date": "supplier_invoice_date",
    "supplier invoice date": "supplier_invoice_date",
    "gateentry_bookno": "gateentry_bookno",
    "gate entry book no": "gateentry_bookno",
    "gate entry no": "gateentry_bookno",
    "gateentry_bookdate": "gateentry_bookdate",
    "gate entry book date": "gateentry_bookdate",
    "tolerance": "tolerance",
    "req_date": "req_date",
    "request date": "req_date",
    "req_person_name": "req_person_name",
    "req person name": "req_person_name",
    "req_person_id": "req_person_id",
    "req person id": "req_person_id",
    "req_department": "req_department",
    "req department": "req_department",
    "req_reason": "req_reason",
    "request reason": "req_reason",
    "supplier_id": "supplier_id",
    "supplier id": "supplier_id",
    "gstin": "gstin",
    "contact_name": "contact_name",
    "contact name": "contact_name",
    "trade_name": "trade_name",
    "trade name": "trade_name",
    "contact_type": "contact_type",
    "contact type": "contact_type",
    "address1": "address1",
    "address 1": "address1",
    "address2": "address2",
    "address 2": "address2",
    "location": "location",
    "city": "location",
    "pincode": "pincode",
    "postal code": "pincode",
    "pin code": "pincode",
    "state_name": "state_name",
    "state name": "state_name",
    "state_code": "state_code",
    "state code": "state_code",
    "country": "country",
    "person_name": "person_name",
    "person name": "person_name",
    "phone_number": "phone_number",
    "phone number": "phone_number",
    "phone": "phone_number",
    "email": "email",
    "category": "category",
    "segment": "segment",
    "sub_segment": "sub_segment",
    "sub segment": "sub_segment",
    "sales_contact_id": "sales_contact_id",
    "sales contact id": "sales_contact_id",
    "currency": "currency",
    "item_id": "item_id",
    "item id": "item_id",
    "item_serial_number": "item_serial_number",
    "item serial number": "item_serial_number",
    "product_description": "product_description",
    "product description": "product_description",
    "hsn_code": "hsn_code",
    "hsn code": "hsn_code",
    "total_quantity": "total_quantity",
    "total quantity": "total_quantity",
    "quantity": "quantity",
    "free_quantity": "free_quantity",
    "free quantity": "free_quantity",
    "accepted_qty": "accepted_qty",
    "accepted qty": "accepted_qty",
    "rejected_qty": "rejected_qty",
    "rejected qty": "rejected_qty",
    "unit": "unit",
    "unit_price": "unit_price",
    "unit price": "unit_price",
    "total_amount": "total_amount",
    "total amount": "total_amount",
    "discount": "discount",
    "assessable_value": "assessable_value",
    "assessable value": "assessable_value",
    "gst_rate": "gst_rate",
    "gst rate": "gst_rate",
    "igst_amount": "igst_amount",
    "igst amount": "igst_amount",
    "cgst_amount": "cgst_amount",
    "cgst amount": "cgst_amount",
    "sgst_amount": "sgst_amount",
    "sgst amount": "sgst_amount",
    "total_item_value": "total_item_value",
    "total item value": "total_item_value",
    "freight_charge": "freight_charge",
    "freight charge": "freight_charge",
    "loading_unloading_charge": "loading_unloading_charge",
    "loading unloading charge": "loading_unloading_charge",
    "total_before_tax": "total_before_tax",
    "total before tax": "total_before_tax",
    "total_tax_amount": "total_tax_amount",
    "total tax amount": "total_tax_amount",
    "total_after_tax": "total_after_tax",
    "total after tax": "total_after_tax",
    "status": "status",
}

MODEL_DATE_FIELDS = {
    "po_date",
    "grn_date",
    "supplier_invoice_date",
    "gateentry_bookdate",
}

MODEL_DATETIME_FIELDS = {
    "delivery_note_date",
    "delivery_note_date",
}

STRING_DATE_FIELDS = {"req_date"}

INTEGER_EDITABLE_FIELDS = {
    "delivery_days_gap",
    "order_rating",
}

DECIMAL_FIELDS = {
    "total_quantity",
    "quantity",
    "free_quantity",
    "accepted_qty",
    "rejected_qty",
    "unit_price",
    "total_amount",
    "assessable_value",
    "gst_rate",
    "igst_amount",
    "cgst_amount",
    "sgst_amount",
    "total_item_value",
    "freight_charge",
    "total_before_tax",
    "total_tax_amount",
    "total_after_tax",
}

INTEGER_FIELDS = {"item_serial_number"}
BOOLEAN_FIELDS = {"status"}
REQUIRED_FIELDS = ("grn_no",)
RECEIVER_NESTED_KEYS = {"document_details", "document_requirement_details", "supplier_details", "items", "value_details"}
ALL_NESTED_KEYS = RECEIVER_NESTED_KEYS | {"invoice_details"}
EDITABLE_DOCUMENT_DETAILS_FIELDS = {
    "gateentry_bookno",
    "gateentry_bookdate",
    "tolerance",
    "supplier_invoice_no",
    "supplier_invoice_date",
}
EDITABLE_REQUIREMENT_DETAILS_FIELDS = {
    "req_date",
    "req_person_name",
    "req_person_id",
    "req_department",
    "req_reason",
}
EDITABLE_INVOICE_DETAILS_FIELDS = {
    "dc_numbers",
    "delivery_days_gap",
    "delivery_note_no",
    "delivery_note_date",
    "order_rating",
    "grn_warehouse",
    "source_warehouse",
    "accepted_warehouse",
    "rejected_warehouse",
}
EDITABLE_ITEM_FIELDS = {
    "item_serial_number",
    "free_quantity",
    "accepted_qty",
    "rejected_qty",
}
EDITABLE_NESTED_FIELD_MAP = {
    "document_details": EDITABLE_DOCUMENT_DETAILS_FIELDS,
    "document_requirement_details": EDITABLE_REQUIREMENT_DETAILS_FIELDS,
    "invoice_details": EDITABLE_INVOICE_DETAILS_FIELDS,
    "items": EDITABLE_ITEM_FIELDS,
}
EDITABLE_FLAT_FIELDS = {
    "gateentry_bookno",
    "gateentry_bookdate",
    "tolerance",
    "supplier_invoice_no",
    "supplier_invoice_date",
    "req_date",
    "req_person_name",
    "req_person_id",
    "req_department",
    "req_reason",
    "item_serial_number",
    "free_quantity",
    "accepted_qty",
    "rejected_qty",
    "dc_numbers",
    "delivery_days_gap",
    "delivery_note_no",
    "delivery_note_date",
    "order_rating",
    "grn_warehouse",
    "source_warehouse",
    "accepted_warehouse",
    "rejected_warehouse",
}

MOVE_TO_QCR_REQUIRED_FIELDS = {
    "gateentry_bookno": "Gate Entry Book No",
    "gateentry_bookdate": "Gate Entry Book Date",
    "grn_warehouse": "Warehouse",
    "accepted_warehouse": "Accepted Warehouse",
    "rejected_warehouse": "Rejected Warehouse",
}
GRN_EDIT_REQUIRED_DOCUMENT_FIELDS = {
    "gateentry_bookno": "Gate Entry Book No",
    "gateentry_bookdate": "Gate Entry Book Date",
}
MANUAL_GATE_ENTRY_FLAG = "_manual_gate_entry_entry"
GATE_ENTRY_STATUS = GRN.PROCESS_STATUS_GATE_ENTRY
GRN_PENDING_STATUS = GRN.PROCESS_STATUS_GRN_PENDING
MOVED_TO_QCR_STATUS = GRN.PROCESS_STATUS_QCR
MOVED_TO_GRN_STATUS = GRN.PROCESS_STATUS_MOVED_TO_GRN
GRN_APPROVED_STATUS = GRN.PROCESS_STATUS_APPROVED
GRN_REJECTED_STATUS = GRN.PROCESS_STATUS_REJECTED
RECEIVER_REQUIRED_FIELDS = (
    "document_details.grn_no",
    "document_details.po_no",
    "document_details.grn_date",
    "supplier_details.supplier_id",
)
RECEIVER_VALIDATION_RESPONSE_CODE = "WPE-VAL-000400"
RECEIVER_INTERNAL_RESPONSE_CODE = "WPE-ERR-000500"
GRN_LEGACY_VALUE_FIELDS = (
    "id",
    "po_no",
    "po_date",
    "grn_no",
    "grn_date",
    "supplier_invoice_no",
    "supplier_invoice_date",
    "gateentry_bookno",
    "gateentry_bookdate",
    "tolerance",
    "req_date",
    "req_person_name",
    "req_person_id",
    "req_department",
    "req_reason",
    "supplier_id",
    "gstin",
    "contact_name",
    "trade_name",
    "contact_type",
    "address1",
    "address2",
    "location",
    "pincode",
    "state_name",
    "state_code",
    "country",
    "person_name",
    "phone_number",
    "email",
    "category",
    "segment",
    "sub_segment",
    "sales_contact_id",
    "currency",
    "item_id",
    "item_serial_number",
    "product_description",
    "hsn_code",
    "total_quantity",
    "quantity",
    "free_quantity",
    "accepted_qty",
    "rejected_qty",
    "unit",
    "unit_price",
    "total_amount",
    "discount",
    "assessable_value",
    "gst_rate",
    "igst_amount",
    "cgst_amount",
    "sgst_amount",
    "total_item_value",
    "freight_charge",
    "loading_unloading_charge",
    "total_before_tax",
    "total_tax_amount",
    "total_after_tax",
    "created_at",
    "updated_at",
    "status",
)


ACTIVE_QCR_SCOPES = {"", "active", "qcr", "pending", "open"}
CANCELLED_QCR_SCOPES = {"cancelled", "canceled", "cancel", "rejected", "reject"}
MOVED_TO_GRN_SCOPES = {"moved to grn", "moved_to_grn", "moved-grn", "grn", "approved", "grn approved", "grn_approved"}
GRN_PENDING_SCOPES = {"pending", "grn pending", "grn_pending", "pending-grn", "pending_grn"}


def resolve_list_scope(request, fallback: str = "") -> str:
    for value in (
        fallback,
        request.query_params.get("tab"),
        request.query_params.get("status"),
        request.query_params.get("view"),
        request.query_params.get("type"),
        request.query_params.get("page"),
        request.query_params.get("scope"),
    ):
        if value is None:
            continue
        normalized = str(value).strip().lower()
        if normalized:
            return normalized
    return ""


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def is_blank_row(row: tuple[Any, ...]) -> bool:
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str) and not cell.strip():
            continue
        return False
    return True


def stringify_cell(value: Any, *, allow_blank: bool = False) -> str | None:
    if value is None:
        return None if allow_blank else ""

    if isinstance(value, bool):
        text = "true" if value else "false"
    elif isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        text = str(int(value)) if value.is_integer() else str(value)
    else:
        text = str(value).strip()

    if not text:
        return None if allow_blank else ""

    return text


def parse_boolean(value: Any, *, default: bool = True) -> bool:
    if value is None or value == "":
        return default

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(int(value))

    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "on", "active", "enabled"}:
        return True
    if normalized in {"0", "false", "no", "n", "off", "inactive", "disabled"}:
        return False

    raise ValueError(f"Cannot interpret boolean value: {value}")


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, bool):
        return Decimal("1") if value else Decimal("0")

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    normalized = str(value).replace(",", "").strip()
    if not normalized:
        return None

    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Cannot interpret decimal value: {value}") from exc


def parse_integer(value: Any) -> int | None:
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value)

    normalized = str(value).replace(",", "").strip()
    if not normalized:
        return None

    try:
        return int(float(normalized))
    except ValueError as exc:
        raise ValueError(f"Cannot interpret integer value: {value}") from exc


def parse_date(value: Any, *, epoch: Any) -> date | None:
    if value is None or value == "":
        return None

    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch=epoch)
        return converted.date() if isinstance(converted, datetime) else converted

    normalized = str(value).strip()
    if not normalized:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Cannot interpret date value: {value}")


def parse_string_date(value: Any, *, epoch: Any) -> str | None:
    parsed_date = parse_date(value, epoch=epoch)
    return parsed_date.isoformat() if parsed_date else None


def format_serializer_errors(errors: dict[str, Any]) -> str:
    parts: list[str] = []

    for field, messages in errors.items():
        if isinstance(messages, (list, tuple)):
            message_text = ", ".join(str(message) for message in messages)
        else:
            message_text = str(messages)
        parts.append(f"{field}: {message_text}")

    return "; ".join(parts) if parts else "Invalid GRN row"


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def get_nested_value(data: dict[str, Any], path: str) -> Any:
    current: Any = data
    for part in path.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


def parse_api_date(value: Any) -> date | None:
    if is_blank(value):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    normalized = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Cannot interpret date value: {value}")


def clean_model_value(field_name: str, value: Any, *, required: bool = False) -> Any:
    if field_name in MODEL_DATE_FIELDS:
        try:
            return parse_api_date(value)
        except ValueError:
            if required:
                raise
            return None

    if field_name in MODEL_DATETIME_FIELDS:
        if is_blank(value):
            return None
        if isinstance(value, str):
            from datetime import datetime as dt
            for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                try:
                    return dt.strptime(value.strip(), fmt)
                except ValueError:
                    continue
        return value

    if field_name in DECIMAL_FIELDS:
        try:
            return parse_decimal(value)
        except ValueError:
            return None

    if field_name in INTEGER_FIELDS or field_name in INTEGER_EDITABLE_FIELDS:
        try:
            return parse_integer(value)
        except ValueError:
            return None

    return value


def validate_grn_receiver_payload(data: Any) -> dict[str, Any]:
    errors: dict[str, Any] = {}

    if not isinstance(data, dict):
        return {"payload": "Expected a JSON object."}

    document_details = data.get("document_details")
    supplier_details = data.get("supplier_details")
    items = data.get("items")

    if not isinstance(document_details, dict):
        errors["document_details"] = "This object is required."
    if not isinstance(supplier_details, dict):
        errors["supplier_details"] = "This object is required."

    for field_path in RECEIVER_REQUIRED_FIELDS:
        if is_blank(get_nested_value(data, field_path)):
            errors[field_path] = "This field is required."

    if not isinstance(items, list) or not items:
        errors["items"] = "At least one item is required."
    elif not all(isinstance(item, dict) for item in items):
        errors["items"] = "Each item must be an object."

    grn_date = get_nested_value(data, "document_details.grn_date")
    if not is_blank(grn_date):
        try:
            parse_api_date(grn_date)
        except ValueError as exc:
            errors["document_details.grn_date"] = str(exc)

    return errors


def map_nested_grn_payload(data: dict[str, Any]) -> dict[str, Any]:
    document_details = data.get("document_details", {}) or {}
    document_requirement_details = data.get("document_requirement_details", {}) or {}
    supplier_details = data.get("supplier_details", {}) or {}
    items = data.get("items", []) or []
    value_details = data.get("value_details", {}) or {}
    invoice_details = data.get("invoice_details", {}) or {}

    first_item = items[0] if items else {}

    return {
        "po_no": document_details.get("po_no"),
        "po_date": document_details.get("po_date"),
        "grn_no": document_details.get("grn_no"),
        "grn_date": document_details.get("grn_date"),
        "supplier_invoice_no": document_details.get("supplier_invoice_no") or invoice_details.get("supplier_invoice_no"),
        "supplier_invoice_date": document_details.get("supplier_invoice_date") or invoice_details.get("supplier_invoice_date"),
        "gateentry_bookno": document_details.get("gateentry_bookno"),
        "gateentry_bookdate": document_details.get("gateentry_bookdate"),
        "tolerance": document_details.get("tolerance"),
        "dc_numbers": invoice_details.get("dc_numbers"),
        "delivery_days_gap": invoice_details.get("delivery_days_gap"),
        "delivery_note_no": invoice_details.get("delivery_note_no"),
        "delivery_note_date": invoice_details.get("delivery_note_date"),
        "order_rating": invoice_details.get("order_rating"),
        "grn_warehouse": invoice_details.get("grn_warehouse"),
        "source_warehouse": invoice_details.get("source_warehouse"),
        "accepted_warehouse": invoice_details.get("accepted_warehouse"),
        "rejected_warehouse": invoice_details.get("rejected_warehouse"),
        "req_date": document_requirement_details.get("req_date"),
        "req_person_name": document_requirement_details.get("req_person_name"),
        "req_person_id": document_requirement_details.get("req_person_id"),
        "req_department": document_requirement_details.get("req_department"),
        "req_reason": document_requirement_details.get("req_reason"),
        "supplier_id": supplier_details.get("supplier_id"),
        "gstin": supplier_details.get("gstin"),
        "contact_name": supplier_details.get("contact_name"),
        "trade_name": supplier_details.get("trade_name"),
        "contact_type": supplier_details.get("contact_type"),
        "address1": supplier_details.get("address1"),
        "address2": supplier_details.get("address2"),
        "location": supplier_details.get("location"),
        "pincode": supplier_details.get("pincode"),
        "state_name": supplier_details.get("state_name"),
        "state_code": supplier_details.get("state_code"),
        "country": supplier_details.get("country"),
        "person_name": supplier_details.get("person_name"),
        "phone_number": supplier_details.get("phone_number"),
        "email": supplier_details.get("email"),
        "category": supplier_details.get("category"),
        "segment": supplier_details.get("segment"),
        "sub_segment": supplier_details.get("sub_segment"),
        "sales_contact_id": supplier_details.get("sales_contact_id"),
        "currency": supplier_details.get("currency"),
        "item_id": first_item.get("item_id"),
        "item_serial_number": first_item.get("item_serial_number"),
        "product_description": first_item.get("product_description"),
        "hsn_code": first_item.get("hsn_code"),
        "total_quantity": first_item.get("total_quantity"),
        "quantity": first_item.get("quantity"),
        "free_quantity": first_item.get("free_quantity"),
        "accepted_qty": first_item.get("accepted_qty"),
        "rejected_qty": first_item.get("rejected_qty"),
        "unit": first_item.get("unit"),
        "unit_price": first_item.get("unit_price"),
        "total_amount": first_item.get("total_amount"),
        "discount": first_item.get("discount"),
        "assessable_value": first_item.get("assessable_value"),
        "gst_rate": first_item.get("gst_rate"),
        "igst_amount": first_item.get("igst_amount"),
        "cgst_amount": first_item.get("cgst_amount"),
        "sgst_amount": first_item.get("sgst_amount"),
        "total_item_value": first_item.get("total_item_value"),
        "freight_charge": value_details.get("freight_charge"),
        "loading_unloading_charge": value_details.get("loading_unloading_charge"),
        "total_before_tax": value_details.get("total_before_tax"),
        "total_tax_amount": value_details.get("total_tax_amount"),
        "total_after_tax": value_details.get("total_after_tax"),
    }


def build_receiver_grn_payload(data: dict[str, Any]) -> dict[str, Any]:
    payload = map_nested_grn_payload(data)

    for field_name, value in list(payload.items()):
        payload[field_name] = clean_model_value(
            field_name,
            value,
            required=field_name == "grn_date",
        )

    for field_name in ("grn_warehouse", "accepted_warehouse", "rejected_warehouse"):
        if payload.get(field_name) in (None, ""):
            payload.pop(field_name, None)

    payload["raw_payload"] = deepcopy(data)
    return payload


def build_grn_edit_payload(grn: GRN) -> dict[str, Any]:
    payload = deepcopy(grn.raw_payload if isinstance(grn.raw_payload, dict) else {})
    serialized = GRNReadSerializer(grn).data

    for section_name in ALL_NESTED_KEYS:
        default_value = deepcopy(serialized.get(section_name))

        if section_name == "items":
            current_items = payload.get(section_name)
            if isinstance(current_items, list):
                normalized_items = [deepcopy(item) if isinstance(item, dict) else {} for item in current_items]
                if not normalized_items and isinstance(default_value, list):
                    normalized_items = default_value
            else:
                normalized_items = default_value if isinstance(default_value, list) else []
            payload[section_name] = normalized_items
            continue

        current_section = payload.get(section_name)
        if not isinstance(current_section, dict):
            current_section = {}

        if isinstance(default_value, dict):
            for field_name, field_value in default_value.items():
                current_section.setdefault(field_name, field_value)

        payload[section_name] = current_section

    return payload


def merge_grn_update_payload(grn: GRN, data: Any) -> dict[str, Any]:
    if not isinstance(data, dict):
        raise DRFValidationError({"payload": "Expected a JSON object."})

    payload = build_grn_edit_payload(grn)

    for section_name, editable_fields in EDITABLE_NESTED_FIELD_MAP.items():
        if section_name == "items":
            continue

        section_updates = data.get(section_name)
        if section_updates is None:
            continue
        if not isinstance(section_updates, dict):
            raise DRFValidationError({section_name: "This object must be provided as a JSON object."})

        target_section = payload.get(section_name)
        if not isinstance(target_section, dict):
            target_section = {}
            payload[section_name] = target_section

        for field_name in editable_fields:
            if field_name in section_updates:
                target_section[field_name] = section_updates.get(field_name)

        if section_name == "document_details" and any(field_name in section_updates for field_name in GRN_EDIT_REQUIRED_DOCUMENT_FIELDS):
            target_section[MANUAL_GATE_ENTRY_FLAG] = True

    item_updates = data.get("items")
    if item_updates is not None:
        if not isinstance(item_updates, list):
            raise DRFValidationError({"items": "This field must be provided as a list."})
        if not all(isinstance(item, dict) for item in item_updates):
            raise DRFValidationError({"items": "Each item must be an object."})

        existing_items = payload.get("items")
        normalized_items = [deepcopy(item) if isinstance(item, dict) else {} for item in existing_items] if isinstance(existing_items, list) else []
        target_length = max(len(normalized_items), len(item_updates))
        merged_items: list[dict[str, Any]] = []

        for index in range(target_length):
            merged_item = deepcopy(normalized_items[index]) if index < len(normalized_items) else {}
            section_updates = item_updates[index] if index < len(item_updates) else {}

            if isinstance(section_updates, dict):
                for field_name in EDITABLE_ITEM_FIELDS:
                    if field_name in section_updates:
                        merged_item[field_name] = section_updates.get(field_name)

            merged_items.append(merged_item)

        payload["items"] = merged_items

    return payload


def to_json_safe(value: Any) -> Any:
    """Recursively convert non-JSON-serializable Python objects to safe primitives."""
    if isinstance(value, dict):
        return {k: to_json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [to_json_safe(v) for v in value]
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def build_grn_update_fields(payload: dict[str, Any]) -> dict[str, Any]:
    flat_payload = map_nested_grn_payload(payload)
    update_payload = {"raw_payload": to_json_safe(payload)}

    for field_name in EDITABLE_FLAT_FIELDS:
        update_payload[field_name] = clean_model_value(field_name, flat_payload.get(field_name))

    return update_payload


def validate_grn_update_payload(payload: dict[str, Any]) -> dict[str, Any]:
    document_details = payload.get("document_details")
    if not isinstance(document_details, dict):
        return {"document_details": "This object is required."}

    errors: dict[str, list[str]] = {}

    for field_name, label in GRN_EDIT_REQUIRED_DOCUMENT_FIELDS.items():
        value = document_details.get(field_name)
        if is_blank(value):
            errors[field_name] = [f"{label} is required."]
            continue

        if field_name == "gateentry_bookdate":
            try:
                parse_api_date(value)
            except ValueError:
                errors[field_name] = [f"{label} must be a valid date."]

    return {"document_details": errors} if errors else {}


def receiver_error_key(field_name: str | None) -> str:
    if is_blank(field_name):
        return "non_field_errors"

    normalized = str(field_name).strip()
    if normalized == "non_field_errors":
        return normalized

    parts = [part for part in normalized.split(".") if part and not part.isdigit()]
    if not parts:
        return "non_field_errors"

    return parts[-1]


def merge_receiver_errors(target: dict[str, list[str]], source: dict[str, list[str]]) -> dict[str, list[str]]:
    for field_name, messages in source.items():
        bucket = target.setdefault(field_name, [])
        for message in messages:
            message_text = str(message)
            if message_text not in bucket:
                bucket.append(message_text)
    return target


def normalize_receiver_errors(errors: Any, *, field_name: str | None = None) -> dict[str, list[str]]:
    normalized: dict[str, list[str]] = {}

    if errors in (None, ""):
        return normalized

    if isinstance(errors, dict):
        for key, value in errors.items():
            next_field_name = key if field_name is None else f"{field_name}.{key}"
            if key == "non_field_errors":
                next_field_name = "non_field_errors"
            merge_receiver_errors(
                normalized,
                normalize_receiver_errors(value, field_name=next_field_name),
            )
        return normalized

    if isinstance(errors, (list, tuple)):
        has_nested_values = any(isinstance(item, (dict, list, tuple)) for item in errors)
        if has_nested_values:
            for item in errors:
                merge_receiver_errors(
                    normalized,
                    normalize_receiver_errors(item, field_name=field_name),
                )
            return normalized

        messages = [str(item) for item in errors if str(item).strip()]
        if messages:
            normalized[receiver_error_key(field_name)] = messages
        return normalized

    normalized[receiver_error_key(field_name)] = [str(errors)]
    return normalized


def format_receiver_timestamp(value: datetime | None = None) -> str:
    moment = value or timezone.now()
    if timezone.is_naive(moment):
        moment = timezone.make_aware(moment, timezone.get_current_timezone())
    return timezone.localtime(moment).isoformat()


def get_actor_name(request) -> str | None:
    user = getattr(request, "user", None)
    if user is not None and getattr(user, "is_authenticated", False):
        full_name = getattr(user, "get_full_name", lambda: "")()
        username = getattr(user, "get_username", lambda: "")()
        return full_name or username or None

    header_actor = request.headers.get("X-User-Name")
    if header_actor:
        return header_actor.strip() or None

    request_actor = request.data.get("moved_by") if hasattr(request.data, "get") else None
    if request_actor:
        return str(request_actor).strip() or None

    return None


def serialize_grn_snapshot(grn: GRN) -> dict[str, Any]:
    snapshot = model_to_dict(grn, exclude=["id", "created_at", "updated_at"])
    snapshot["id"] = grn.id
    snapshot["created_at"] = grn.created_at.isoformat() if grn.created_at else None
    snapshot["updated_at"] = grn.updated_at.isoformat() if grn.updated_at else None

    for field_name, value in list(snapshot.items()):
        if isinstance(value, Decimal):
            snapshot[field_name] = str(value)
        elif isinstance(value, (date, datetime)):
            snapshot[field_name] = value.isoformat()

    return snapshot


def build_store_sync_payload(*, grn: GRN, qcr_status: str | None = None, accepted: bool = True) -> dict[str, Any]:
    payload = deepcopy(grn.raw_payload if isinstance(grn.raw_payload, dict) else {})
    document_details = payload.get("document_details")
    if not isinstance(document_details, dict):
        document_details = {}
        payload["document_details"] = document_details
    supplier_details = payload.get("supplier_details")
    if not isinstance(supplier_details, dict):
        supplier_details = {}
        payload["supplier_details"] = supplier_details

    document_details.setdefault("grn_no", grn.grn_no)
    if grn.grn_date:
        document_details.setdefault("grn_date", grn.grn_date.isoformat())
    supplier_details.setdefault("trade_name", grn.trade_name)

    raw_items = payload.get("items")
    item_lines = [item for item in raw_items if isinstance(item, dict)] if isinstance(raw_items, list) else []
    if not item_lines:
        item_lines = [{}]
        payload["items"] = item_lines

    first_item = item_lines[0]
    first_item.setdefault("item_id", grn.item_id)
    first_item.setdefault("product_description", grn.product_description)
    first_item.setdefault("hsn_code", grn.hsn_code)
    first_item.setdefault("unit", grn.unit)
    if grn.total_quantity is not None:
        first_item.setdefault("total_quantity", str(grn.total_quantity))
    if grn.quantity is not None:
        first_item.setdefault("quantity", str(grn.quantity))
    if grn.accepted_qty is not None:
        first_item.setdefault("accepted_qty", str(grn.accepted_qty))

    payload["id"] = grn.id
    payload["unique_id"] = grn.unique_id
    payload["grn_no"] = grn.grn_no
    payload.setdefault("item_id", grn.item_id)
    payload.setdefault("product_description", grn.product_description)
    payload.setdefault("unit", grn.unit)
    payload.setdefault("trade_name", grn.trade_name)
    payload["process_status"] = grn.process_status
    if qcr_status:
        payload["qcr_status"] = qcr_status
    payload["target_warehouse"] = grn.accepted_warehouse or "Stores" if accepted else grn.rejected_warehouse or "Rejected Warehouse - CBE"
    if not accepted:
        payload["use_rejected_qty"] = True
    if grn.grn_date:
        payload.setdefault("grn_date", grn.grn_date.isoformat())
    if grn.accepted_qty is not None:
        payload.setdefault("accepted_qty", str(grn.accepted_qty))
    if grn.quantity is not None:
        payload.setdefault("quantity", str(grn.quantity))
    if grn.total_quantity is not None:
        payload.setdefault("total_quantity", str(grn.total_quantity))
    return payload


def is_missing_schema_error(exc: Exception) -> bool:
    if not isinstance(exc, (ProgrammingError, OperationalError)):
        return False

    message = str(exc).lower()
    return "doesn't exist" in message or "unknown column" in message or "no such column" in message


def schema_sync_response() -> Response:
    return Response(
        {
            "status": "error",
            "message": "Database schema is not up to date. Apply the latest Purchases_Inwards migration before using the GRN to QCR flow.",
            "code": "schema_out_of_sync",
        },
        status=status.HTTP_503_SERVICE_UNAVAILABLE,
    )


def get_compatible_grn_payload() -> list[dict[str, Any]]:
    rows = list(GRN.objects.values(*GRN_LEGACY_VALUE_FIELDS).order_by("-id"))
    for row in rows:
        row["department"] = row.get("req_department")
        row["process_status"] = GATE_ENTRY_STATUS
        row["moved_to_qcr_at"] = None
        row["moved_to_qcr_by"] = None
    return rows


class GRNAPIViewMixin:
    def get_base_queryset(self, request):
        queryset = GRN.objects.all()
        list_scope = resolve_list_scope(request, getattr(self, "tab_scope", ""))

        if list_scope in MOVED_TO_GRN_SCOPES:
            return queryset.filter(process_status__in=[MOVED_TO_GRN_STATUS, GRN_APPROVED_STATUS])
        if list_scope in GRN_PENDING_SCOPES:
            return queryset.filter(process_status=GRN_PENDING_STATUS)
        return queryset.filter(process_status=GATE_ENTRY_STATUS)

    def get_grn_response(self, request):
        try:
            queryset = self.get_base_queryset(request).order_by("-id")

            grn_id = request.query_params.get("id")
            grn_no = request.query_params.get("grn_no")

            if grn_id:
                queryset = queryset.filter(id=grn_id)
            if grn_no:
                queryset = queryset.filter(grn_no=grn_no)

            if (grn_id or grn_no) and not queryset.exists():
                return Response(
                    {
                        "status": "error",
                        "message": "GRN not found",
                        "count": 0,
                        "data": [],
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            serializer = GRNReadSerializer(queryset, many=True)
            return Response(
                {
                    "status": "success",
                    "message": "GRN data fetched successfully",
                    "count": queryset.count(),
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                try:
                    return Response(get_compatible_grn_payload(), status=status.HTTP_200_OK)
                except (ProgrammingError, OperationalError):
                    return schema_sync_response()
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception as exc:
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def create_grn_response(self, request):
        try:
            data = request.data
            if any(key in data for key in RECEIVER_NESTED_KEYS):
                payload = build_receiver_grn_payload(data)
            elif hasattr(data, "dict"):
                payload = data.dict()
            else:
                payload = dict(data)

            serializer = GRNSerializer(data=payload)
            if serializer.is_valid():
                saved_grn = serializer.save()
                GRNAuditLog.objects.create(
                    grn=saved_grn,
                    stage=GRNAuditLog.STAGE_GRN_CREATED,
                    actor=get_actor_name(request),
                    notes=f"GRN {saved_grn.grn_no} created.",
                )
                return Response(
                    {
                        "status": "success",
                        "message": "GRN stored successfully",
                        "grn_no": saved_grn.grn_no,
                        "grn_id": saved_grn.id,
                    },
                    status=status.HTTP_201_CREATED,
                )

            return Response(
                {
                    "status": "error",
                    "message": "Validation failed",
                    "errors": serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                return schema_sync_response()
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception as exc:
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GRNCreateAPIView(GRNAPIViewMixin, APIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]
    tab_scope = ""

    def get(self, request):
        return self.get_grn_response(request)

    def post(self, request):
        return self.create_grn_response(request)


class GRNDetailAPIView(APIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get(self, request, pk: int):
        try:
            grn = GRN.objects.filter(pk=pk).first()
            if grn is None:
                return Response(
                    {
                        "status": "error",
                        "message": "GRN not found",
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            return Response(
                {
                    "status": "success",
                    "message": "GRN detail fetched successfully",
                    "data": GRNReadSerializer(grn).data,
                },
                status=status.HTTP_200_OK,
            )
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                return schema_sync_response()
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception as exc:
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def patch(self, request, pk: int):
        try:
            grn = GRN.objects.filter(pk=pk).first()
            if grn is None:
                return Response(
                    {
                        "status": "error",
                        "message": "GRN not found",
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            if grn.process_status != GATE_ENTRY_STATUS:
                return Response(
                    {
                        "status": "error",
                        "message": "GRN details can only be edited in Gate Entry stage. This record is locked.",
                        "process_status": grn.process_status,
                    },
                    status=status.HTTP_409_CONFLICT,
                )

            merged_payload = merge_grn_update_payload(grn, request.data)
            payload_errors = validate_grn_update_payload(merged_payload)
            if payload_errors:
                return Response(
                    {
                        "status": "error",
                        "message": "Validation failed",
                        "errors": payload_errors,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            serializer = GRNSerializer(
                grn,
                data=build_grn_update_fields(merged_payload),
                partial=True,
            )
            if not serializer.is_valid():
                return Response(
                    {
                        "status": "error",
                        "message": "Validation failed",
                        "errors": serializer.errors,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            saved_grn = serializer.save()
            GRNAuditLog.objects.create(
                grn=saved_grn,
                stage=GRNAuditLog.STAGE_GRN_EDITED,
                actor=get_actor_name(request),
                notes="Gate Entry fields updated.",
            )
            return Response(
                {
                    "status": "success",
                    "message": "GRN updated successfully",
                    "data": GRNReadSerializer(saved_grn).data,
                },
                status=status.HTTP_200_OK,
            )
        except DRFValidationError as exc:
            return Response(
                {
                    "status": "error",
                    "message": "Validation failed",
                    "errors": exc.detail,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                return schema_sync_response()
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception as exc:
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GRNReceiverCreateAPIView(APIView):
    authentication_classes = []
    parser_classes = [JSONParser]
    permission_classes = [AllowAny]
    renderer_classes = [JSONRenderer]

    def build_response(
        self,
        *,
        http_status: int,
        status_text: str,
        message: str,
        response_code: str,
        grn_no: str | None,
        receiver_reference: str | None,
        received_at: datetime | None = None,
        errors: dict[str, list[str]] | None,
    ) -> Response:
        return Response(
            {
                "status": status_text,
                "message": message,
                "response_code": response_code,
                "grn_no": grn_no,
                "receiver_reference": receiver_reference,
                "received_at": format_receiver_timestamp(received_at),
                "errors": errors,
            },
            status=http_status,
        )

    def build_grn_response_code(self, grn: GRN) -> str:
        return f"WPE-GRN-{grn.pk:06d}"

    def normalize_text(self, value: Any) -> str | None:
        if is_blank(value):
            return None
        return str(value).strip()

    def extract_request_grn_no(self, data: Any, *, fallback: str | None = None) -> str | None:
        grn_no = fallback
        if isinstance(data, dict):
            grn_no = get_nested_value(data, "document_details.grn_no") or data.get("grn_no") or fallback
        return self.normalize_text(grn_no)

    def find_existing_grn(self, *, grn_no: str | None, idempotency_key: str | None) -> GRN | None:
        lookup_values: list[str] = []
        for candidate in (grn_no, idempotency_key):
            normalized = self.normalize_text(candidate)
            if normalized and normalized not in lookup_values:
                lookup_values.append(normalized)

        if not lookup_values:
            return None

        return GRN.objects.filter(grn_no__in=lookup_values).order_by("id").first()

    def build_success_response(self, grn: GRN) -> Response:
        return self.build_response(
            http_status=status.HTTP_201_CREATED,
            status_text="sent",
            message="GRN received and processed successfully.",
            response_code=self.build_grn_response_code(grn),
            grn_no=grn.grn_no,
            receiver_reference=grn.unique_id,
            received_at=grn.created_at,
            errors=None,
        )

    def build_duplicate_response(self, grn: GRN, *, grn_no: str | None = None) -> Response:
        return self.build_response(
            http_status=status.HTTP_200_OK,
            status_text="duplicate",
            message="GRN already exists.",
            response_code=self.build_grn_response_code(grn),
            grn_no=grn_no or grn.grn_no,
            receiver_reference=grn.unique_id,
            received_at=grn.created_at,
            errors=None,
        )

    def build_validation_error_response(self, *, grn_no: str | None, errors: Any) -> Response:
        return self.build_response(
            http_status=status.HTTP_400_BAD_REQUEST,
            status_text="error",
            message="Payload validation failed.",
            response_code=RECEIVER_VALIDATION_RESPONSE_CODE,
            grn_no=grn_no,
            receiver_reference=None,
            errors=normalize_receiver_errors(errors),
        )

    def build_internal_error_response(self, *, grn_no: str | None) -> Response:
        return self.build_response(
            http_status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            status_text="error",
            message="Internal processing failed.",
            response_code=RECEIVER_INTERNAL_RESPONSE_CODE,
            grn_no=grn_no,
            receiver_reference=None,
            errors={"non_field_errors": ["Unexpected server error."]},
        )

    def post(self, request):
        idempotency_key = self.normalize_text(request.headers.get("Idempotency-Key"))
        grn_no = idempotency_key

        try:
            data = request.data
            grn_no = self.extract_request_grn_no(data, fallback=idempotency_key)

            existing_grn = self.find_existing_grn(grn_no=grn_no, idempotency_key=idempotency_key)
            if existing_grn is not None:
                return self.build_duplicate_response(existing_grn, grn_no=grn_no)

            errors = validate_grn_receiver_payload(data)
            if errors:
                return self.build_validation_error_response(grn_no=grn_no, errors=errors)

            payload = build_receiver_grn_payload(data)
            payload["grn_no"] = grn_no
            serializer = GRNSerializer(data=payload)
            if not serializer.is_valid():
                existing_grn = self.find_existing_grn(grn_no=grn_no, idempotency_key=idempotency_key)
                if existing_grn is not None:
                    return self.build_duplicate_response(existing_grn, grn_no=grn_no)
                return self.build_validation_error_response(grn_no=grn_no, errors=serializer.errors)

            try:
                with transaction.atomic():
                    saved_grn = serializer.save()
            except IntegrityError:
                existing_grn = self.find_existing_grn(grn_no=grn_no, idempotency_key=idempotency_key)
                if existing_grn is not None:
                    return self.build_duplicate_response(existing_grn, grn_no=grn_no)
                raise
        except DRFValidationError as exc:
            return self.build_validation_error_response(grn_no=grn_no, errors=exc.detail)
        except (ParseError, UnsupportedMediaType) as exc:
            return self.build_validation_error_response(
                grn_no=grn_no,
                errors={"payload": exc.detail},
            )
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                return self.build_internal_error_response(grn_no=grn_no)
            return self.build_internal_error_response(grn_no=grn_no)
        except Exception:
            return self.build_internal_error_response(grn_no=grn_no)

        return self.build_success_response(saved_grn)


class GRNViewSet(GRNAPIViewMixin, viewsets.ViewSet):
    def list(self, request):
        return self.get_grn_response(request)

    def create(self, request):
        return self.create_grn_response(request)


class GRNImportAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if uploaded_file is None:
            return Response(
                {"detail": "Please upload an Excel file using the 'file' field."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not uploaded_file.name.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            return Response(
                {"detail": "Only .xlsx Excel files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
        except Exception:
            return Response(
                {"detail": "The uploaded file is not a valid Excel workbook."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            sheet = workbook.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return Response(
                    {"detail": "The workbook does not contain a header row."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            column_map: dict[int, str] = {}
            for index, header in enumerate(header_row):
                field_name = HEADER_ALIASES.get(normalize_header(header))
                if field_name and field_name not in column_map.values():
                    column_map[index] = field_name

            missing_required_fields = [field for field in REQUIRED_FIELDS if field not in column_map.values()]
            if missing_required_fields:
                return Response(
                    {"detail": "The workbook is missing required columns: " + ", ".join(missing_required_fields)},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            created_count = 0
            failed_rows: list[dict[str, Any]] = []
            processed_count = 0

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if is_blank_row(row):
                    continue

                processed_count += 1
                payload: dict[str, Any] = {}

                try:
                    for column_index, field_name in column_map.items():
                        cell_value = row[column_index] if column_index < len(row) else None

                        if field_name in MODEL_DATE_FIELDS:
                            payload[field_name] = parse_date(cell_value, epoch=workbook.epoch)
                        elif field_name in STRING_DATE_FIELDS:
                            payload[field_name] = parse_string_date(cell_value, epoch=workbook.epoch)
                        elif field_name in DECIMAL_FIELDS:
                            payload[field_name] = parse_decimal(cell_value)
                        elif field_name in INTEGER_FIELDS:
                            payload[field_name] = parse_integer(cell_value)
                        elif field_name in BOOLEAN_FIELDS:
                            payload[field_name] = parse_boolean(cell_value, default=True)
                        else:
                            payload[field_name] = stringify_cell(cell_value, allow_blank=True)
                except ValueError as exc:
                    failed_rows.append({"row": row_number, "message": str(exc)})
                    continue

                serializer = GRNSerializer(data=payload)
                if not serializer.is_valid():
                    failed_rows.append(
                        {
                            "row": row_number,
                            "message": format_serializer_errors(serializer.errors),
                            "details": serializer.errors,
                        }
                    )
                    continue

                try:
                    with transaction.atomic():
                        serializer.save()
                except IntegrityError as exc:
                    failed_rows.append({"row": row_number, "message": str(exc)})
                    continue

                created_count += 1

            failed_count = len(failed_rows)
            response_body = {
                "created_count": created_count,
                "failed_count": failed_count,
                "processed_count": processed_count,
                "errors": failed_rows,
            }

            if created_count == 0:
                if failed_rows:
                    response_body["detail"] = (
                        f"All rows failed to import. First error: row {failed_rows[0]['row']}: "
                        f"{failed_rows[0]['message']}"
                    )
                else:
                    response_body["detail"] = "The workbook does not contain any GRN rows."
                return Response(response_body, status=status.HTTP_400_BAD_REQUEST)

            if failed_count > 0:
                return Response(response_body, status=status.HTTP_207_MULTI_STATUS)

            return Response(response_body, status=status.HTTP_201_CREATED)
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                return schema_sync_response()
            raise
        finally:
            workbook.close()


class QCRListAPIView(APIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]
    tab_scope = ""

    def get(self, request):
        try:
            list_scope = resolve_list_scope(request, self.tab_scope)

            queryset = QCR.objects.select_related("source_grn")
            if list_scope in CANCELLED_QCR_SCOPES:
                queryset = queryset.filter(
                    Q(status=GRN_REJECTED_STATUS)
                    | Q(status=MOVED_TO_GRN_STATUS, source_grn__qc_status="Partial")
                )
            elif list_scope in MOVED_TO_GRN_SCOPES:
                queryset = queryset.filter(status=MOVED_TO_GRN_STATUS)
            elif list_scope == "all":
                queryset = queryset.all()
            else:
                queryset = queryset.filter(status="Active")

            queryset = queryset.order_by("-id")
            serializer = QCRSerializer(queryset, many=True)
            return Response(serializer.data)
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                return Response([])
            raise


class WarehouseInventoryListAPIView(APIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get(self, request):
        warehouse_name = (request.query_params.get("warehouse_name") or "").strip()
        if not warehouse_name:
            return Response(
                {"warehouse_name": "The 'warehouse_name' query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            rows = build_warehouse_inventory_rows(warehouse_name)
            search_term = (request.query_params.get("search") or "").strip()
            if search_term:
                rows = [row for row in rows if warehouse_inventory_row_matches_search(row, search_term)]
            return Response(paginate_warehouse_inventory_rows(request, rows))
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                return Response(
                    {
                        "count": 0,
                        "next": None,
                        "previous": None,
                        "results": [],
                    }
                )
            raise


def validate_move_to_grn_pending(grn: GRN) -> dict[str, str]:
    errors: dict[str, str] = {}
    for field_name, label in MOVE_TO_QCR_REQUIRED_FIELDS.items():
        value = getattr(grn, field_name, None)
        if is_blank(value):
            errors[field_name] = f"{label} is required before moving to GRN Pending."

    raw_payload = grn.raw_payload if isinstance(grn.raw_payload, dict) else {}
    document_details = raw_payload.get("document_details")
    has_manual_gate_entry = isinstance(document_details, dict) and document_details.get(MANUAL_GATE_ENTRY_FLAG) is True
    if not has_manual_gate_entry:
        for field_name, label in GRN_EDIT_REQUIRED_DOCUMENT_FIELDS.items():
            if field_name not in errors:
                errors[field_name] = f"{label} must be entered manually before moving to GRN Pending."
    return errors


def decimal_to_string(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(value, "f")


def parse_optional_decimal(value: Any, *, default: Decimal = Decimal("0")) -> Decimal:
    text = "" if value is None else str(value).strip()
    if not text:
        return default
    try:
        return Decimal(text)
    except (InvalidOperation, TypeError, ValueError):
        return default


def normalize_warehouse_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def warehouse_labels_match(left: Any, right: Any) -> bool:
    left_label = normalize_warehouse_label(left)
    right_label = normalize_warehouse_label(right)
    if not left_label or not right_label:
        return False
    return left_label == right_label


def resolve_warehouse_inventory_scope(warehouse_name: str) -> str:
    normalized = normalize_warehouse_label(warehouse_name)
    if "qc pending" in normalized:
        return "qc_pending"
    if "reject" in normalized:
        return "rejected"
    return "generic"


def dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        normalized = value.strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        result.append(normalized)
    return result


def build_warehouse_inventory_rows(warehouse_name: str) -> list[dict[str, Any]]:
    scope = resolve_warehouse_inventory_scope(warehouse_name)
    rows: list[dict[str, Any]] = []

    queryset = QCR.objects.select_related("source_grn").order_by("-id")
    for qcr_record in queryset:
        grn = qcr_record.source_grn
        pending_items = [item for item in (grn.grn_pending_items or []) if isinstance(item, dict)]
        if not pending_items:
            continue

        matching_items: list[dict[str, Any]] = []
        inward_qty = Decimal("0")
        outward_qty = Decimal("0")
        reasons: list[str] = []

        if scope == "qc_pending":
            matching_items = [
                item
                for item in pending_items
                if warehouse_labels_match(item.get("store_in_name") or item.get("store_in"), warehouse_name)
            ]
            if not matching_items:
                continue
            inward_qty = sum((parse_optional_decimal(item.get("received_qty")) for item in matching_items), Decimal("0"))
            if qcr_record.status != "Active":
                outward_qty = inward_qty
        elif scope == "rejected":
            if qcr_record.status == "Active":
                continue
            if not warehouse_labels_match(grn.rejected_warehouse or "Rejected Warehouse - CBE", warehouse_name):
                continue
            matching_items = [
                item for item in pending_items if parse_optional_decimal(item.get("rejected_qty")) > Decimal("0")
            ]
            if not matching_items:
                continue
            inward_qty = sum((parse_optional_decimal(item.get("rejected_qty")) for item in matching_items), Decimal("0"))
            reasons = dedupe_preserve_order(
                [
                    str(item.get("rejection_reason") or "").strip()
                    for item in matching_items
                    if str(item.get("rejection_reason") or "").strip()
                ]
            )
        else:
            matching_items = [
                item
                for item in pending_items
                if warehouse_labels_match(item.get("store_in_name") or item.get("store_in"), warehouse_name)
            ]
            if not matching_items:
                continue
            inward_qty = sum((parse_optional_decimal(item.get("received_qty")) for item in matching_items), Decimal("0"))
            if qcr_record.status != "Active":
                outward_qty = inward_qty

        item_names = dedupe_preserve_order(
            [
                str(
                    item.get("item_name")
                    or item.get("product_description")
                    or item.get("item_id")
                    or ""
                )
                for item in matching_items
            ]
        )
        supplier_name = (
            grn.trade_name
            or (grn.raw_payload.get("supplier_details", {}) if isinstance(grn.raw_payload, dict) else {}).get("trade_name")
            or ""
        )
        po_number = (
            grn.po_no
            or (grn.raw_payload.get("document_details", {}) if isinstance(grn.raw_payload, dict) else {}).get("po_no")
            or ""
        )

        rows.append(
            {
                "grn_id": grn.id,
                "qcr_id": qcr_record.id,
                "grn_no": grn.grn_no,
                "supplier": supplier_name,
                "po_no": po_number,
                "items": ", ".join(item_names),
                "inward_qty": decimal_to_string(inward_qty) or "0",
                "outward_qty": decimal_to_string(outward_qty) or "0",
                "status": grn.process_status,
                "reason": " | ".join(reasons) if reasons else (qcr_record.remarks or ""),
                "warehouse_name": warehouse_name,
                "moved_to_qcr_at": grn.moved_to_qcr_at.isoformat() if grn.moved_to_qcr_at else None,
            }
        )

    return rows


def warehouse_inventory_row_matches_search(row: dict[str, Any], search_term: str) -> bool:
    needle = normalize_warehouse_label(search_term)
    if not needle:
        return True

    haystacks = (
        row.get("grn_no"),
        row.get("supplier"),
        row.get("po_no"),
        row.get("items"),
        row.get("status"),
        row.get("inward_qty"),
        row.get("outward_qty"),
        row.get("reason"),
    )
    return any(needle in normalize_warehouse_label(value) for value in haystacks if value not in (None, ""))


def parse_positive_int_param(value: Any, *, default: int, minimum: int = 1, maximum: int | None = None) -> int:
    try:
        parsed = int(str(value).strip()) if value not in (None, "") else default
    except (TypeError, ValueError):
        return default

    if parsed < minimum:
        parsed = minimum
    if maximum is not None and parsed > maximum:
        parsed = maximum
    return parsed


def build_paginated_link(request, *, page_number: int, page_size: int) -> str:
    query_params = request.query_params.copy()
    query_params["page"] = str(page_number)
    query_params["page_size"] = str(page_size)
    query_string = urlencode([(key, value) for key, values in query_params.lists() for value in values])
    return request.build_absolute_uri(f"{request.path}?{query_string}")


def paginate_warehouse_inventory_rows(request, rows: list[dict[str, Any]]) -> dict[str, Any]:
    page = parse_positive_int_param(request.query_params.get("page"), default=1)
    page_size = parse_positive_int_param(request.query_params.get("page_size"), default=20, maximum=200)
    total = len(rows)
    start = (page - 1) * page_size
    end = start + page_size

    next_link = None
    if end < total:
        next_link = build_paginated_link(request, page_number=page + 1, page_size=page_size)

    previous_link = None
    if page > 1:
        previous_link = build_paginated_link(request, page_number=page - 1, page_size=page_size)

    return {
        "count": total,
        "next": next_link,
        "previous": previous_link,
        "results": rows[start:end],
    }


def parse_decimal_value(value: Any, *, field_label: str) -> Decimal:
    text = "" if value is None else str(value).strip()
    if not text:
        raise DRFValidationError({field_label: f"{field_label} is required."})
    try:
        decimal_value = Decimal(text)
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise DRFValidationError({field_label: f"{field_label} must be numeric."}) from exc

    if decimal_value < 0:
        raise DRFValidationError({field_label: f"{field_label} cannot be negative."})
    return decimal_value


def parse_decimal_value_or_default(
    value: Any,
    *,
    field_label: str,
    default: Decimal = Decimal("0"),
) -> Decimal:
    text = "" if value is None else str(value).strip()
    if not text:
        return default
    return parse_decimal_value(text, field_label=field_label)


def ensure_grn_payload_item_lines(grn: GRN, payload: dict[str, Any] | None = None) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    payload_data = deepcopy(payload if isinstance(payload, dict) else grn.raw_payload if isinstance(grn.raw_payload, dict) else {})
    raw_items = payload_data.get("items")
    item_lines = [deepcopy(item) for item in raw_items if isinstance(item, dict)] if isinstance(raw_items, list) else []

    if not item_lines:
        item_lines = [{}]
        payload_data["items"] = item_lines

    first_item = item_lines[0]
    first_item.setdefault("item_id", grn.item_id)
    first_item.setdefault("item_serial_number", grn.item_serial_number)
    first_item.setdefault("product_description", grn.product_description)
    first_item.setdefault("hsn_code", grn.hsn_code)
    first_item.setdefault("unit", grn.unit)
    if grn.total_quantity is not None:
        first_item.setdefault("total_quantity", str(grn.total_quantity))
    if grn.quantity is not None:
        first_item.setdefault("quantity", str(grn.quantity))
    if grn.accepted_qty is not None:
        first_item.setdefault("accepted_qty", str(grn.accepted_qty))
    if grn.rejected_qty is not None:
        first_item.setdefault("rejected_qty", str(grn.rejected_qty))

    payload_data["items"] = item_lines
    return payload_data, item_lines


def resolve_sent_qty(item_line: dict[str, Any]) -> Decimal | None:
    for field_name in ("quantity", "total_quantity", "accepted_qty"):
        value = item_line.get(field_name)
        if value in (None, ""):
            continue
        try:
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            continue
    return None


def build_grn_pending_items(grn: GRN, item_lines: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    lines = item_lines if item_lines is not None else ensure_grn_payload_item_lines(grn)[1]
    pending_items: list[dict[str, Any]] = []
    for index, item_line in enumerate(lines):
        sent_qty = resolve_sent_qty(item_line)
        pending_items.append(
            {
                "line_index": index,
                "item_id": item_line.get("item_id"),
                "item_name": item_line.get("product_description") or item_line.get("item_id"),
                "unit": item_line.get("unit"),
                "sent_qty": decimal_to_string(sent_qty) if sent_qty is not None else item_line.get("quantity") or item_line.get("total_quantity"),
                "received_qty": item_line.get("received_qty") or item_line.get("accepted_qty"),
                "store_in_id": item_line.get("store_in_id") or item_line.get("store_in"),
                "store_in_name": item_line.get("store_in_name") or item_line.get("store_in"),
                "accepted_qty": item_line.get("accepted_qty"),
                "rejected_qty": item_line.get("rejected_qty"),
                "rejection_reason": item_line.get("rejection_reason") or "",
                "pending_rejected_qty": item_line.get("pending_rejected_qty"),
            }
        )
    return pending_items


def resolve_qcr_received_qty(item_line: dict[str, Any], pending_item: dict[str, Any], *, index: int) -> Decimal:
    for candidate in (
        pending_item.get("received_qty"),
        item_line.get("received_qty"),
        item_line.get("accepted_qty"),
        item_line.get("quantity"),
        item_line.get("total_quantity"),
    ):
        if candidate in (None, ""):
            continue
        try:
            decimal_value = Decimal(str(candidate))
        except (InvalidOperation, TypeError, ValueError):
            continue
        if decimal_value < 0:
            break
        return decimal_value
    raise DRFValidationError({f"items[{index}].received_qty": "Received Qty is not available for this QCR line."})


def build_qcr_completion_remarks(qcr_items: list[dict[str, Any]]) -> str | None:
    remark_lines = []
    for item in qcr_items:
        if item.get("rejected_qty") in (None, "", "0", "0.0", "0.00", "0.000"):
            continue
        reason = str(item.get("rejection_reason") or "").strip()
        if not reason:
            continue
        item_name = item.get("item_name") or item.get("item_id") or f"Line {int(item.get('line_index', 0)) + 1}"
        remark_lines.append(f"{item_name}: {reason}")
    if not remark_lines:
        return None
    return " | ".join(remark_lines)


def normalize_qcr_completion_items(
    qcr_record: QCR,
    payload_items: Any,
    *,
    actor: str | None,
    completed_at: datetime,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any], Decimal, Decimal, str | None]:
    if not isinstance(payload_items, list) or not payload_items:
        raise DRFValidationError({"items": "At least one item row is required to complete this QCR."})

    grn = qcr_record.source_grn
    raw_payload, item_lines = ensure_grn_payload_item_lines(grn)
    stored_pending_items = grn.grn_pending_items if isinstance(grn.grn_pending_items, list) else []
    pending_items = stored_pending_items if len(stored_pending_items) == len(item_lines) else build_grn_pending_items(grn, item_lines)

    if len(payload_items) != len(item_lines):
        raise DRFValidationError({"items": "All QCR item rows must be completed before submission."})

    payload_items_by_index: dict[int, dict[str, Any]] = {}
    for index, payload_item in enumerate(payload_items):
        if not isinstance(payload_item, dict):
            continue
        raw_line_index = payload_item.get("line_index")
        try:
            line_index = int(raw_line_index) if raw_line_index not in (None, "") else index
        except (TypeError, ValueError):
            line_index = index
        payload_items_by_index[line_index] = payload_item

    total_accepted = Decimal("0")
    total_rejected = Decimal("0")
    completed_at_text = completed_at.isoformat()
    normalized_qcr_items: list[dict[str, Any]] = []
    normalized_pending_items: list[dict[str, Any]] = []

    for index, item_line in enumerate(item_lines):
        pending_item = pending_items[index] if index < len(pending_items) and isinstance(pending_items[index], dict) else {}
        raw_line_index = pending_item.get("line_index", index)
        try:
            line_index = int(raw_line_index)
        except (TypeError, ValueError):
            line_index = index
        item_payload = payload_items_by_index.get(line_index)
        if item_payload is None:
            raise DRFValidationError({f"items[{index}]": "This QCR line is missing from the submission payload."})

        received_qty = resolve_qcr_received_qty(item_line, pending_item, index=index)
        rejected_qty = parse_decimal_value_or_default(
            item_payload.get("rejected_qty"),
            field_label=f"items[{index}].rejected_qty",
        )
        if rejected_qty > received_qty:
            raise DRFValidationError({f"items[{index}].rejected_qty": "Rejected Qty cannot exceed Received Qty."})

        accepted_qty = received_qty - rejected_qty
        reason_text = str(item_payload.get("reason") or item_payload.get("rejection_reason") or "").strip()
        if rejected_qty > 0 and not reason_text:
            raise DRFValidationError({f"items[{index}].reason": "Reason is required when Rejected Qty is greater than zero."})

        sent_qty = resolve_sent_qty(item_line)
        pending_rejected_qty = Decimal("0")
        if sent_qty is not None and sent_qty > received_qty:
            pending_rejected_qty = sent_qty - received_qty

        total_accepted += accepted_qty
        total_rejected += rejected_qty

        item_name = (
            pending_item.get("item_name")
            or item_line.get("product_description")
            or item_line.get("item_name")
            or item_line.get("item_id")
            or f"Line {index + 1}"
        )
        store_in_id = pending_item.get("store_in_id") or item_line.get("store_in_id") or item_line.get("store_in")
        store_in_name = pending_item.get("store_in_name") or item_line.get("store_in_name") or item_line.get("store_in")
        unit = pending_item.get("unit") or item_line.get("unit")
        sent_qty_text = decimal_to_string(sent_qty) if sent_qty is not None else item_line.get("quantity") or item_line.get("total_quantity")

        item_line["received_qty"] = decimal_to_string(received_qty)
        item_line["accepted_qty"] = decimal_to_string(accepted_qty)
        item_line["rejected_qty"] = decimal_to_string(rejected_qty)
        item_line["rejection_reason"] = reason_text
        item_line["pending_rejected_qty"] = decimal_to_string(pending_rejected_qty)
        item_line["qcr_completed_at"] = completed_at_text
        item_line["qcr_completed_by"] = actor

        normalized_item = {
            "line_index": line_index,
            "item_id": item_line.get("item_id"),
            "item_name": item_name,
            "unit": unit,
            "sent_qty": sent_qty_text,
            "received_qty": decimal_to_string(received_qty),
            "accepted_qty": decimal_to_string(accepted_qty),
            "rejected_qty": decimal_to_string(rejected_qty),
            "rejection_reason": reason_text,
            "pending_rejected_qty": decimal_to_string(pending_rejected_qty),
            "store_in_id": store_in_id,
            "store_in_name": store_in_name or str(store_in_id or ""),
        }
        normalized_qcr_items.append(normalized_item)
        normalized_pending_items.append(dict(normalized_item))

    raw_payload["items"] = item_lines
    raw_payload["qcr_items"] = normalized_qcr_items
    raw_payload["qcr_completed_at"] = completed_at_text
    raw_payload["qcr_completed_by"] = actor
    return (
        normalized_qcr_items,
        normalized_pending_items,
        raw_payload,
        total_accepted,
        total_rejected,
        build_qcr_completion_remarks(normalized_qcr_items),
    )


def normalize_pending_to_qcr_items(grn: GRN, payload_items: Any) -> tuple[list[dict[str, Any]], dict[str, Any], Decimal, Decimal]:
    if not isinstance(payload_items, list) or not payload_items:
        raise DRFValidationError({"items": "At least one item row is required to move this GRN to QCR."})

    raw_payload, item_lines = ensure_grn_payload_item_lines(grn)
    if len(payload_items) != len(item_lines):
        raise DRFValidationError({"items": "All GRN item rows must be completed before moving to QCR."})

    total_received = Decimal("0")
    total_rejected = Decimal("0")
    normalized_items: list[dict[str, Any]] = []

    for index, item_line in enumerate(item_lines):
        item_payload = payload_items[index] if isinstance(payload_items[index], dict) else {}
        sent_qty = resolve_sent_qty(item_line)
        received_qty = parse_decimal_value(item_payload.get("received_qty"), field_label=f"items[{index}].received_qty")
        store_in_id = item_payload.get("store_in_id") or item_payload.get("store_in")
        store_in_name = item_payload.get("store_in_name") or item_payload.get("store_in_name_label") or item_payload.get("store_in")

        if is_blank(store_in_id) and is_blank(store_in_name):
            raise DRFValidationError({f"items[{index}].store_in": "Store In is required."})

        if sent_qty is not None and received_qty > sent_qty:
            raise DRFValidationError({f"items[{index}].received_qty": "Received Qty cannot exceed Sent Qty."})

        rejected_qty = Decimal("0") if sent_qty is None else sent_qty - received_qty
        total_received += received_qty
        total_rejected += rejected_qty

        item_line["received_qty"] = decimal_to_string(received_qty)
        item_line["accepted_qty"] = decimal_to_string(received_qty)
        item_line["rejected_qty"] = decimal_to_string(rejected_qty)
        item_line["store_in_id"] = store_in_id
        item_line["store_in_name"] = store_in_name or str(store_in_id)
        item_line["store_in"] = store_in_name or str(store_in_id)

        normalized_items.append(
            {
                "line_index": index,
                "item_id": item_line.get("item_id"),
                "item_name": item_line.get("product_description") or item_line.get("item_id"),
                "unit": item_line.get("unit"),
                "sent_qty": decimal_to_string(sent_qty) if sent_qty is not None else item_line.get("quantity") or item_line.get("total_quantity"),
                "received_qty": decimal_to_string(received_qty),
                "rejected_qty": decimal_to_string(rejected_qty),
                "store_in_id": store_in_id,
                "store_in_name": store_in_name or str(store_in_id),
            }
        )

    raw_payload["items"] = item_lines
    return normalized_items, raw_payload, total_received, total_rejected


class GRNMoveToQCRAPIView(APIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def post(self, request, pk: int):
        try:
            with transaction.atomic():
                grn = GRN.objects.select_for_update().filter(pk=pk).first()
                if grn is None:
                    return Response(
                        {
                            "status": "error",
                            "message": "The selected GRN record was not found.",
                        },
                        status=status.HTTP_404_NOT_FOUND,
                    )

                if grn.process_status != GATE_ENTRY_STATUS:
                    return Response(
                        {
                            "status": "error",
                            "message": "Only Gate Entry records can be moved to GRN Pending.",
                        },
                        status=status.HTTP_409_CONFLICT,
                    )

                validation_errors = validate_move_to_grn_pending(grn)
                if validation_errors:
                    return Response(
                        {
                            "status": "error",
                            "message": "Mandatory fields are missing. Please complete all required fields before moving to GRN Pending.",
                            "errors": validation_errors,
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                item_lines = ensure_grn_payload_item_lines(grn)[1]
                moved_by = get_actor_name(request)
                grn.process_status = GRN_PENDING_STATUS
                grn.qc_status = "Pending"
                grn.status = True
                grn.grn_pending_items = build_grn_pending_items(grn, item_lines)
                grn.save(update_fields=["process_status", "qc_status", "status", "grn_pending_items", "updated_at"])

                GRNAuditLog.objects.create(
                    grn=grn,
                    stage=GRNAuditLog.STAGE_MOVED_TO_GRN_PENDING,
                    actor=moved_by,
                    notes="Moved from Gate Entry to GRN Pending.",
                )

            return Response(
                {
                    "status": "success",
                    "message": "GRN record moved to GRN Pending successfully.",
                    "grn": GRNSerializer(grn).data,
                },
                status=status.HTTP_200_OK,
            )
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                return schema_sync_response()
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GRNPendingToQCRAPIView(APIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def post(self, request, pk: int):
        try:
            with transaction.atomic():
                grn = GRN.objects.select_for_update().filter(pk=pk).first()
                if grn is None:
                    return Response(
                        {
                            "status": "error",
                            "message": "The selected GRN record was not found.",
                        },
                        status=status.HTTP_404_NOT_FOUND,
                    )

                if grn.process_status != GRN_PENDING_STATUS:
                    return Response(
                        {
                            "status": "error",
                            "message": "Only GRN Pending records can be moved to QCR.",
                        },
                        status=status.HTTP_409_CONFLICT,
                    )

                if QCR.objects.filter(source_grn=grn).exists():
                    return Response(
                        {
                            "status": "error",
                            "message": "This GRN record has already been moved to QCR.",
                        },
                        status=status.HTTP_409_CONFLICT,
                    )

                normalized_items, raw_payload, total_received, total_rejected = normalize_pending_to_qcr_items(
                    grn,
                    request.data.get("items") if hasattr(request.data, "get") else None,
                )

                moved_at = timezone.now()
                moved_by = get_actor_name(request)

                grn.raw_payload = raw_payload
                grn.grn_pending_items = normalized_items
                grn.accepted_qty = total_received
                grn.rejected_qty = total_rejected
                grn.process_status = MOVED_TO_QCR_STATUS
                grn.qc_status = "Pending"
                grn.status = False
                grn.moved_to_qcr_at = moved_at
                grn.moved_to_qcr_by = moved_by
                grn.save(
                    update_fields=[
                        "raw_payload",
                        "grn_pending_items",
                        "accepted_qty",
                        "rejected_qty",
                        "process_status",
                        "qc_status",
                        "status",
                        "moved_to_qcr_at",
                        "moved_to_qcr_by",
                        "updated_at",
                    ]
                )

                snapshot = serialize_grn_snapshot(grn)
                qcr_record = QCR.objects.create(
                    source_grn=grn,
                    grn_reference_no=grn.grn_no,
                    snapshot=snapshot,
                    status="Active",
                    moved_to_qcr_at=moved_at,
                    moved_to_qcr_by=moved_by,
                )

                GRNAuditLog.objects.create(
                    grn=grn,
                    stage=GRNAuditLog.STAGE_MOVED_TO_QCR,
                    actor=moved_by,
                    notes=f"Moved to QCR from GRN Pending. QCR ID: {qcr_record.unique_id}",
                )

            return Response(
                {
                    "status": "success",
                    "message": "GRN record moved to QCR successfully.",
                    "grn": GRNSerializer(grn).data,
                    "qcr": QCRSerializer(qcr_record).data,
                },
                status=status.HTTP_201_CREATED,
            )
        except DRFValidationError as exc:
            return Response(
                {
                    "status": "error",
                    "message": "Validation failed",
                    "errors": exc.detail,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                return schema_sync_response()
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception as exc:
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception as exc:
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class QCRStatusUpdateAPIView(APIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def post(self, request, pk: int):
        action = request.data.get("action") if hasattr(request.data, "get") else None
        if action not in {"complete", "move_to_grn", "reject"}:
            return Response(
                {
                    "status": "error",
                    "message": "A valid QCR action is required (complete, move_to_grn, or reject).",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        remarks = None
        if action == "reject":
            raw_remarks = request.data.get("remarks") if hasattr(request.data, "get") else None
            remarks = str(raw_remarks).strip() if raw_remarks else None
            if not remarks:
                return Response(
                    {
                        "status": "error",
                        "message": "Remarks are mandatory when rejecting a QCR.",
                        "errors": {"remarks": "This field is required for rejection."},
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        try:
            with transaction.atomic():
                qcr_record = QCR.objects.select_for_update().select_related("source_grn").filter(pk=pk).first()
                if qcr_record is None:
                    return Response(
                        {
                            "status": "error",
                            "message": "The selected QCR record was not found.",
                        },
                        status=status.HTTP_404_NOT_FOUND,
                    )

                if qcr_record.status != "Active":
                    return Response(
                        {
                            "status": "error",
                            "message": "This QCR record has already been processed.",
                        },
                        status=status.HTTP_409_CONFLICT,
                    )

                grn = qcr_record.source_grn
                actor = get_actor_name(request)

                if action == "complete":
                    completed_at = timezone.now()
                    (
                        normalized_qcr_items,
                        normalized_pending_items,
                        raw_payload,
                        total_accepted,
                        total_rejected,
                        completion_remarks,
                    ) = normalize_qcr_completion_items(
                        qcr_record,
                        request.data.get("items") if hasattr(request.data, "get") else None,
                        actor=actor,
                        completed_at=completed_at,
                    )

                    qcr_record.qcr_items = normalized_qcr_items
                    qcr_record.qcr_completed_at = completed_at
                    qcr_record.qcr_completed_by = actor
                    qcr_record.remarks = completion_remarks

                    grn.raw_payload = raw_payload
                    grn.grn_pending_items = normalized_pending_items
                    grn.accepted_qty = total_accepted
                    grn.rejected_qty = total_rejected

                    if total_accepted > 0:
                        qcr_record.status = MOVED_TO_GRN_STATUS
                        grn.process_status = GRN_APPROVED_STATUS
                        grn.qc_status = "Pass" if total_rejected == 0 else "Partial"
                        grn.status = True
                        message = (
                            "QCR completed. Accepted quantity moved to store."
                            if total_rejected == 0
                            else "QCR completed. Only accepted quantity moved to store."
                        )
                        audit_stage = GRNAuditLog.STAGE_QCR_ACCEPTED
                        audit_notes = (
                            f"QC completed with accepted quantity {decimal_to_string(total_accepted)} moved to "
                            f"{grn.accepted_warehouse or 'Stores'}."
                        )
                        if total_rejected > 0:
                            audit_notes = (
                                f"{audit_notes} Rejected quantity: {decimal_to_string(total_rejected)}. "
                                f"Reasons: {completion_remarks or 'Line-level rejection reasons captured.'}"
                            )
                    else:
                        qcr_record.status = GRN_REJECTED_STATUS
                        grn.process_status = GRN_REJECTED_STATUS
                        grn.qc_status = "Fail"
                        grn.status = False
                        if not qcr_record.remarks:
                            qcr_record.remarks = "All received quantity was rejected during QCR."
                        message = "QCR completed. No accepted quantity was moved to store."
                        audit_stage = GRNAuditLog.STAGE_QCR_REJECTED
                        audit_notes = (
                            "QC completed with zero accepted quantity. "
                            f"Rejected quantity: {decimal_to_string(total_rejected)}. "
                            f"Reasons: {qcr_record.remarks}"
                        )

                    qcr_record.save(
                        update_fields=[
                            "qcr_items",
                            "status",
                            "remarks",
                            "qcr_completed_at",
                            "qcr_completed_by",
                            "updated_at",
                        ]
                    )
                    grn.save(
                        update_fields=[
                            "raw_payload",
                            "grn_pending_items",
                            "accepted_qty",
                            "rejected_qty",
                            "process_status",
                            "qc_status",
                            "status",
                            "updated_at",
                        ]
                    )
                elif action == "move_to_grn":
                    qcr_record.status = MOVED_TO_GRN_STATUS
                    qcr_record.qcr_completed_at = timezone.now()
                    qcr_record.qcr_completed_by = actor
                    qcr_record.save(update_fields=["status", "qcr_completed_at", "qcr_completed_by", "updated_at"])
                    grn.process_status = GRN_APPROVED_STATUS
                    grn.qc_status = "Pass"
                    grn.status = True
                    message = "QCR accepted. GRN approved and stock added to store."
                    audit_stage = GRNAuditLog.STAGE_QCR_ACCEPTED
                    audit_notes = f"QC Pass. Accepted quantity moved to {grn.accepted_warehouse or 'Stores'}."
                else:
                    qcr_record.status = GRN_REJECTED_STATUS
                    qcr_record.remarks = remarks
                    qcr_record.qcr_completed_at = timezone.now()
                    qcr_record.qcr_completed_by = actor
                    qcr_record.save(
                        update_fields=["status", "remarks", "qcr_completed_at", "qcr_completed_by", "updated_at"]
                    )
                    grn.process_status = GRN_REJECTED_STATUS
                    grn.qc_status = "Fail"
                    grn.status = False
                    message = "QCR rejected. Stock will not be moved to store."
                    audit_stage = GRNAuditLog.STAGE_QCR_REJECTED
                    audit_notes = f"QC Fail. Remarks: {remarks}. Quantity not moved to store."

                grn.save(update_fields=["process_status", "qc_status", "status", "updated_at"])

                GRNAuditLog.objects.create(
                    grn=grn,
                    stage=audit_stage,
                    actor=actor,
                    notes=audit_notes,
                )

                if action in {"complete", "move_to_grn"} and grn.accepted_qty is not None and grn.accepted_qty > 0:
                    sync_payload = build_store_sync_payload(
                        grn=grn,
                        qcr_status=qcr_record.status,
                        accepted=True,
                    )
                    add_stock_from_grn(
                        sync_payload,
                        created_by=getattr(request, "user", None),
                    )
                    GRNAuditLog.objects.create(
                        grn=grn,
                        stage=GRNAuditLog.STAGE_ADDED_TO_STORE,
                        actor=actor,
                        notes=f"Inventory posted to {grn.accepted_warehouse or 'Stores'}.",
                    )

            return Response(
                {
                    "status": "success",
                    "message": message,
                    "grn": GRNSerializer(grn).data,
                    "qcr": QCRSerializer(qcr_record).data,
                },
                status=status.HTTP_200_OK,
            )
        except DRFValidationError as exc:
            error_message = "Unable to process the QCR request."
            if action == "complete":
                error_message = "QCR could not be completed."
            elif action == "move_to_grn":
                error_message = "Store stock could not be synced from this GRN."
            elif action == "reject":
                error_message = "QCR rejection could not be processed."
            return Response(
                {
                    "status": "error",
                    "message": error_message,
                    "errors": exc.detail,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except (ProgrammingError, OperationalError) as exc:
            if is_missing_schema_error(exc):
                return schema_sync_response()
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception as exc:
            return Response(
                {
                    "status": "error",
                    "message": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class GRNAuditLogAPIView(APIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get(self, request, pk: int):
        try:
            grn = GRN.objects.filter(pk=pk).first()
            if grn is None:
                return Response(
                    {"status": "error", "message": "GRN not found."},
                    status=status.HTTP_404_NOT_FOUND,
                )
            logs = GRNAuditLog.objects.filter(grn=grn).order_by("timestamp")
            serializer = GRNAuditLogSerializer(logs, many=True)
            return Response({"status": "success", "data": serializer.data})
        except Exception as exc:
            return Response(
                {"status": "error", "message": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
