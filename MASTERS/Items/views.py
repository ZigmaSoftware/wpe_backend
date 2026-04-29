import re
from typing import Any

from django.db import IntegrityError, transaction
from openpyxl import load_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response

from .models import Item
from .serializers import ItemSerializer


HEADER_ALIASES = {
    "category": "category",
    "group": "group",
    "sub_group": "sub_group",
    "subgroup": "sub_group",
    "sub group": "sub_group",
    "item_name": "item_name",
    "item name": "item_name",
    "item_code": "item_code",
    "item code": "item_code",
    "hsn_code": "hsn_code",
    "hsn code": "hsn_code",
    "hsn": "hsn_code",
    "unit": "unit",
    "product_details": "product_details",
    "product details": "product_details",
    "description": "description",
    "min_max_status": "min_max_status",
    "min max status": "min_max_status",
    "status": "status",
}

REQUIRED_FIELDS = ("category", "group", "sub_group", "item_name", "item_code", "unit")
BOOLEAN_FIELDS = {"min_max_status", "status"}
TEXT_DEFAULTS = {
    "hsn_code": None,
    "product_details": None,
    "description": None,
}
BOOLEAN_DEFAULTS = {
    "min_max_status": False,
    "status": True,
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""

    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def is_blank_row(row: tuple[Any, ...]) -> bool:
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str) and not cell.strip():
            continue
        return False
    return True


def stringify_cell(value: Any, *, allow_blank: bool = False) -> str | None:
    if value is None:
        return None if allow_blank else ""

    if isinstance(value, bool):
        text = "true" if value else "false"
    elif isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        text = str(int(value)) if value.is_integer() else str(value)
    else:
        text = str(value).strip()

    if not text:
        return None if allow_blank else ""

    return text


def parse_boolean(value: Any, *, default: bool) -> bool:
    if value is None or value == "":
        return default

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(int(value))

    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "on", "active", "enabled"}:
        return True
    if normalized in {"0", "false", "no", "n", "off", "inactive", "disabled"}:
        return False

    raise ValueError(f"Cannot interpret boolean value: {value}")


def format_serializer_errors(errors: dict[str, Any]) -> str:
    parts: list[str] = []

    for field, messages in errors.items():
        if isinstance(messages, (list, tuple)):
            message_text = ", ".join(str(message) for message in messages)
        else:
            message_text = str(messages)
        parts.append(f"{field}: {message_text}")

    return "; ".join(parts) if parts else "Invalid item row"


class ItemViewSet(viewsets.ModelViewSet):
    parser_classes = [JSONParser, FormParser, MultiPartParser]
    queryset = Item.objects.all().order_by('-id')
    serializer_class = ItemSerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        category = self.request.query_params.get('category')
        group = self.request.query_params.get('group')
        sub_group = self.request.query_params.get('sub_group')

        if category:
            queryset = queryset.filter(category=category)
        if group:
            queryset = queryset.filter(group=group)
        if sub_group:
            queryset = queryset.filter(sub_group=sub_group)

        return queryset

    @action(detail=False, methods=["post"], url_path="import")
    def import_excel(self, request):
        uploaded_file = request.FILES.get("file")
        if uploaded_file is None:
            return Response(
                {"detail": "Please upload an Excel file using the 'file' field."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not uploaded_file.name.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            return Response(
                {"detail": "Only .xlsx Excel files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
        except Exception:
            return Response(
                {"detail": "The uploaded file is not a valid Excel workbook."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            sheet = workbook.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return Response(
                    {"detail": "The workbook does not contain a header row."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            column_map: dict[int, str] = {}
            for index, header in enumerate(header_row):
                field_name = HEADER_ALIASES.get(normalize_header(header))
                if field_name and field_name not in column_map.values():
                    column_map[index] = field_name

            missing_required_fields = [field for field in REQUIRED_FIELDS if field not in column_map.values()]
            if missing_required_fields:
                return Response(
                    {
                        "detail": "The workbook is missing required columns: " + ", ".join(missing_required_fields)
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            created_count = 0
            failed_rows: list[dict[str, Any]] = []
            processed_count = 0

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if is_blank_row(row):
                    continue

                processed_count += 1
                payload: dict[str, Any] = {**TEXT_DEFAULTS, **BOOLEAN_DEFAULTS}

                try:
                    for column_index, field_name in column_map.items():
                        cell_value = row[column_index] if column_index < len(row) else None

                        if field_name in BOOLEAN_FIELDS:
                            payload[field_name] = parse_boolean(cell_value, default=BOOLEAN_DEFAULTS[field_name])
                        elif field_name in TEXT_DEFAULTS:
                            payload[field_name] = stringify_cell(cell_value, allow_blank=True)
                        else:
                            payload[field_name] = stringify_cell(cell_value)
                except ValueError as exc:
                    failed_rows.append(
                        {
                            "row": row_number,
                            "message": str(exc),
                        }
                    )
                    continue

                serializer = ItemSerializer(data=payload)
                if not serializer.is_valid():
                    failed_rows.append(
                        {
                            "row": row_number,
                            "message": format_serializer_errors(serializer.errors),
                            "details": serializer.errors,
                        }
                    )
                    continue

                try:
                    with transaction.atomic():
                        serializer.save()
                except IntegrityError as exc:
                    failed_rows.append(
                        {
                            "row": row_number,
                            "message": str(exc),
                        }
                    )
                    continue

                created_count += 1

            failed_count = len(failed_rows)
            response_body = {
                "created_count": created_count,
                "failed_count": failed_count,
                "processed_count": processed_count,
                "errors": failed_rows,
            }

            if created_count == 0:
                if failed_rows:
                    response_body["detail"] = (
                        f"All rows failed to import. First error: row {failed_rows[0]['row']}: "
                        f"{failed_rows[0]['message']}"
                    )
                else:
                    response_body["detail"] = "The workbook does not contain any item rows."
                return Response(response_body, status=status.HTTP_400_BAD_REQUEST)

            if failed_count > 0:
                return Response(response_body, status=status.HTTP_207_MULTI_STATUS)

            return Response(response_body, status=status.HTTP_201_CREATED)
        finally:
            workbook.close()
